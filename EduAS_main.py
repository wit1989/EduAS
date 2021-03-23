#!/usr/bin/env python3
# -*- coding:utf-8 -*-
import cx_Oracle, os, openpyxl, sys, re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, colors, Alignment, Border
import openpyxl.worksheet.protection
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import ttk
import datetime

g_font = ('Monaco', 12)
s_font = ('Monaco', 12)

kb = []
result = []
teachercount = []


# 计算函数
# 绩点计算
def jdjs(cj):
    if cj >= 60:
        jd = (cj - 50) / 10
    else:
        jd = 0
    return jd


# 成绩折算后比较，返回最终绩点
# 将多等级分制改成数字，并比较正考、补考、重修成绩，返回最高绩点
def cj_num_max(list):
    if list[-4] == None:
        zkcj = 0
    elif list[-4] in cjdzb.keys():
        zkcj = cjdzb[list[-4]]
    else:
        zkcj = float(list[-4])

    if list[-3] == None:
        bkcj = 0
    elif list[-3] in cjdzb.keys():
        bkcj = cjdzb[list[-3]]
    else:
        bkcj = float(list[-3])

    bkjd = (lambda b: 1 if b >= 60 else 0)(bkcj)

    if list[-2] == None:
        cxcj = 0
    elif list[-2] in cjdzb.keys():
        cxcj = cjdzb[list[-2]]
    else:
        cxcj = float(list[-2])

    return max(jdjs(zkcj), bkjd, jdjs(cxcj))


# 判断审核指标是否为空
def input_null(inp):
    if inp == '':
        return 0
    else:
        return float(inp)


# 排序函数
def treeview_sort_column(tv, col, reverse):
    """Treeview点击字段排序"""
    l = [(tv.set(k, col), k) for k in tv.get_children('')]
    l.sort(key=lambda t: (t[0]), reverse=reverse)

    for index, (val, k) in enumerate(l):
        tv.move(k, '', index)

    tv.heading(col,
               command=lambda: treeview_sort_column(tv, col, not reverse))


# 1、功能函数

def audit_conn():
    """毕业资格审核-连接数据库"""
    global cjdzb, xsjbxxb, bys_re, aud_crl_sh, aud_crl_sc, zydmb, j, Error_data
    conn = cx_Oracle.connect('zfxfzb/zfsoft_hqwy@orcl')
    cjdzb = {}
    cur = conn.cursor()
    sql_cjdzb = 'SELECT cj,dycj FROM CJDZB'
    cur.execute(sql_cjdzb)
    for cjdzb_row in cur:
        cjdzb[cjdzb_row[0]] = cjdzb_row[1]
    cur.close()

    xsjbxxb = {}
    bys_re = {}

    cur = conn.cursor()
    '''
    创建保存成绩统计信息的字典bys_re {'xh':'[(学生基本信息)],{学生成绩信息}'}
    例子：
    2015002472
    [(2019, '杨思谣', '女', '中东欧语学院', '俄语', '俄语1503', '是', None),
    {'jd': 0, 'gk': 0, 'tyk': 0, 'ggbxk': 0, 'zybxk': 0, 'gxk': 0, 'gxkxf': 0, 'zxk': 0, 'zxkxf': 0, 'zhjnxf': 0, 'sjhj': 0, 'kcxfjd': 0, 'zxf': 0}]


    含不在校生
    GRA_JBXXB: SELECT XH,DQSZJ+XZ BYSJ,XM,XB,XY,ZYMC,XZB FROM XSJBXXB WHERE sfzx='是' and 
              DQSZJ + XZ = (select (case when to_number(to_char(sysdate,'MM')) <=7 then to_number(to_char(sysdate,'YYYY')) 
              else to_number(to_char(sysdate,'YYYY'))+1 end ) from dual) ORDER BY XY,ZYMC,XZB,SFZX,XH
    '''
    sql_xsjbxxb = r"SELECT * FROM GRA_JBXXB"
    cur.execute(sql_xsjbxxb)
    for student in cur:
        tj = []
        tj.append(student[1:9])
        tj.append(
            {'jd': 0, 'gk': 0, 'tyk': 0, 'ggbxk': 0, 'zybxk': 0, 'gxk': 0, 'gxkxf': 0, 'zxk': 0, 'zxkxf': 0, 'bgzdh': 0,
             'zhjnxf': 0, 'sjhj': 0, 'kcxfjd': 0, 'zxf': 0})
        bys_re[student[0]] = tj
    j = cur.execute(sql_xsjbxxb).fetchone()[1]  # 用于获取毕业年份
    cur.close()

    """筛选数据存在问题的学生，放入Erro_data"""
    Error_data = audit_lookforError(conn)

    conn.close()

    xydmb = {}
    xydmbs = []
    for bys_xx in bys_re.values():  # 把学生数据中的学院信息放到list中，然后去重、排序
        xydmbs.append(bys_xx[0][3])
    xydmblist = list(set(xydmbs))  # 去重
    xydmblist.sort(reverse=True)  # 排序
    aud_crl_xyc['values'] = xydmblist

    xy_num = 1  # 生成供用户选择的学院菜单列表
    for bys_xy in xydmblist:
        xydmb[xy_num] = bys_xy
        xy_num += 1

    zydmb = {}
    for no, xy in xydmb.items():
        # 创建专业列表
        zydmb[xy] = []
    for bys_xx in bys_re.values():
        if bys_xx[0][4] not in zydmb[bys_xx[0][3]]:
            zydmb[bys_xx[0][3]].append(bys_xx[0][4])

    aud_crl_conn.configure(text='连接成功')  # 设置button显示的内容
    aud_crl_conn.configure(state='disabled')  # 将按钮设置为灰色状态，不可使用状态
    aud_crl_sh.configure(state='normal')  # 暂时设为灰色，查询成功后显示


def audit_choosezy(ev=None):
    """点击专业选择控件时 按照选择的学院生成专业菜单"""
    global zydmb
    aud_crl_zyc['values'] = zydmb[aud_crl_xyc.get()]


def begin_audit(ev=None):
    """毕业资格审核-开始审核"""
    global aud_crl_xyc, aud_crl_zyc, cjdzb, xsjbxxb, aud_crl_sc, zydmb, xyxz, zymc, aud_crl_sh, var_sh, root, xymc, standard, Error_data, aud_crl_treev, bys_re
    aud_crl_sh.configure(state='disable')
    aud_crl_sc.configure(state='disable')

    """bys_re中的成绩信息归零"""
    for person in bys_re.values():
        for cj_data in person[1].keys():
            person[1][cj_data] = 0

    xymc = []
    zymc = []
    xyxz = 1
    if aud_crl_xyc.get() == '':
        xyxz = 0
        for zy in zydmb.values():
            for i in zy:
                zymc.append(i)
    elif aud_crl_zyc.get() == '':
        zymc = zydmb[aud_crl_xyc.get()]
        xymc.append(aud_crl_xyc.get())
    else:
        zymc.append(aud_crl_zyc.get())
        xymc.append(aud_crl_xyc.get())
    # print(zydmb)
    # print(xymc,zymc)

    standard = {'jd': 2, 'gk': 0, 'tyk': 7, 'ggbxk': 0, 'zybxk': 0, 'gxk': 0, 'gxkxf': 0, 'zxk': 0, 'zxkxf': 0,
                'bgzdh': 1, 'zhjnxf': 0, 'sjhj': 0}
    # standard_keys = ['jd', 'gk', 'tyk', 'ggbxk', 'zybxk', 'gxk', 'gxkxf', 'zxk', 'zxkxf', 'bgzdh', 'zhjnxf', 'sjhj']
    standard['ggbxk'] = aud_crl_gbs.get()
    standard['zybxk'] = aud_crl_zbs.get()
    standard['gxk'] = aud_crl_txs.get()
    standard['gxkxf'] = aud_crl_txxfs.get()
    standard['zxk'] = aud_crl_zxs.get()
    # standard['zxkxf'] = aud_crl_zxxfs.get()
    if aud_crl_zxxfe.get() == "":
        standard['zxkxf'] = 0
    else:
        standard['zxkxf'] = aud_crl_zxxfe.get()
    standard['zhjnxf'] = aud_crl_zjs.get()
    standard['sjhj'] = aud_crl_sjs.get()

    # def begin_audit2(ev=None):
    #     global aud_crl_xyc, aud_crl_zyc, cjdzb, xsjbxxb, bys_re, aud_crl_sc, zydmb, xyxz, zymc, aud_crl_sh, var_sh, root
    num = 1
    load = 0
    conn = cx_Oracle.connect('zfxfzb/zfsoft_hqwy@orcl')
    cur = conn.cursor()
    """
    含不在校生 
    sql_cjb = r"CREATE OR REPLACE VIEW GRA_CJB AS
                SELECT XH,XKKH,KCMC,KCXZ,XF,CJ ZKCJ,BKCJ,CXCJ,CXBJ FROM CJB WHERE XH IN
                (SELECT XH FROM XSJBXXB WHERE sfzx='是' and 
                DQSZJ + XZ = (select (case when to_number(to_char(sysdate,'MM')) <=7 
                then to_number(to_char(sysdate,'YYYY')) 
                else to_number(to_char(sysdate,'YYYY'))+1 end ) from dual)) AND (FXBJ IS NULL OR FXBJ = '0')"
    """
    sql_cjb = r"SELECT * FROM GRA_CJB WHERE KCMC IS NOT NULL"
    cur.execute(sql_cjb)
    for cjb_row in cur:
        if cj_num_max(cjb_row) == 0:
            if cjb_row[3] not in ('公共选修课', '通识选修课'):
                bys_re[cjb_row[0]][1]['gk'] += 1
                if cjb_row[3] in ('公共必修课', '通识必修课', '必修', '英语必修课', '专业必修课', '专业方向必修课',
                                  '专业基础必修课', '专业选修课'):
                    bys_re[cjb_row[0]][1]['zxf'] += float(cjb_row[4])
        if cj_num_max(cjb_row) > 0:
            if re.match(
                    '(搏击操|健康悦跑|健美操|毽球|篮球|轮滑|排球|攀岩|乒乓球|跆拳道|体育|网球|游泳|瑜伽|羽毛球|足球)(俱乐部|训练队)?[1-8]?|体质健康课|田径训练队|武术训练队|康复保健',
                    cjb_row[2]):
                bys_re[cjb_row[0]][1]['tyk'] += 1
            elif cjb_row[2] == '办公自动化':
                bys_re[cjb_row[0]][1]['bgzdh'] += 1
            elif cjb_row[3] == '综合技能训练':
                bys_re[cjb_row[0]][1]['zhjnxf'] += float(cjb_row[4])
            elif cjb_row[3] in ('通识选修课', '公共选修课'):
                bys_re[cjb_row[0]][1]['gxk'] += 1
                bys_re[cjb_row[0]][1]['gxkxf'] += float(cjb_row[4])
            elif cjb_row[2] in ('毕业论文', '毕业实习', '专业实习'):
                bys_re[cjb_row[0]][1]['sjhj'] += 1

            if cjb_row[3] in ('公共必修课', '通识必修课', '必修', '英语必修课'):
                bys_re[cjb_row[0]][1]['ggbxk'] += 1
                bys_re[cjb_row[0]][1]['zxf'] += float(cjb_row[4])
                bys_re[cjb_row[0]][1]['kcxfjd'] += float(cjb_row[4]) * cj_num_max(cjb_row)
            elif cjb_row[3] in ('专业必修课', '专业方向必修课', '专业基础必修课'):
                bys_re[cjb_row[0]][1]['zybxk'] += 1
                bys_re[cjb_row[0]][1]['zxf'] += float(cjb_row[4])
                bys_re[cjb_row[0]][1]['kcxfjd'] += float(cjb_row[4]) * cj_num_max(cjb_row)
            elif cjb_row[3] == '专业选修课':
                bys_re[cjb_row[0]][1]['zxk'] += 1
                bys_re[cjb_row[0]][1]['zxkxf'] += float(cjb_row[4])
                bys_re[cjb_row[0]][1]['zxf'] += float(cjb_row[4])
                bys_re[cjb_row[0]][1]['kcxfjd'] += float(cjb_row[4]) * cj_num_max(cjb_row)
        num += 1
        if int(num / 1600) in range(1, 101):
            if int(num / 1600) != load:
                load = int(num / 1600)
                sys.stdout.write('\r%s%%' % load)
                var_sh.set('正在审核...' + str(load) + "%")

        root.update()
    cur.close()
    conn.close()
    # print(cjb_row)   # ('2015002089', '(2018-2019-1)-16019017-14004-1', '礼仪实训', '综合技能训练', '1', 81.0, None, None)

    # 计算绩点
    for xh, xx in bys_re.items():
        if xx[1]['zxf'] == 0:
            xx[1]['jd'] = 0
        else:
            xx[1]['jd'] = round((xx[1]['kcxfjd'] / xx[1]['zxf']), 2)

    if load == 100:
        var_sh.set('审核完成，点击可再次审核')
        # aud_crl_sh.configure(state='normal')
        aud_crl_sc.configure(state='normal')  # 暂时设为灰色，查询成功后显示

    """将有问题的学生信息根据Input的学院和专业进行筛选，最终得到要显示出来的学号列表Error_data_xh"""
    if xyxz == 0:
        Error_data_xh = [err[0] for err in Error_data]
    else:
        Error_data_xh = []
        for err in Error_data:
            if err[-1] in zymc:
                Error_data_xh.append(err[0])

    """根据Error_data_xh找到学生的全部要显示信息放到Error_data_op中"""

    Error_data_op = []
    for std in Error_data_xh:
        student = []
        student.append(std)
        student.append(bys_re[std][0][1])
        student.append(bys_re[std][0][3])
        student.append(bys_re[std][0][5])
        student.append(bys_re[std][1]['jd'])
        student.append(bys_re[std][1]['gk'])
        Error_data_op.append(student)

    print(Error_data_op)

    # print(Error_data_xh)

    # 清空Treeview中的数据
    x = aud_crl_treev.get_children()
    for item in x:
        aud_crl_treev.delete(item)

    # 将新数据放入Treeview
    for s in Error_data_op:
        aud_crl_treev.insert('', s[0], text=s[0], values=s[0:])

    # for a, b in bys_re.items():
    #     print(a,b)


def audit_toexcel():
    global standard, xymc, xyxz, zymc, j, aud_crl_otext
    wb = openpyxl.Workbook()  # 打开新的工作簿

    wb.create_sheet(index=1, title='全部审核数据')  # 在第二个位置创建新的工作表，并命名为：全部审核数据
    result = wb.worksheets[0]
    result.title = '审核结果'
    alldata = wb.worksheets[1]

    # 写入每列字段名
    col = 1
    for tit in ('序号', '学号', '毕业时间', '姓名', '性别', '学院', '专业名称', '行政班', '绩点',
                '挂科门数(不含公选)', '体育课通过门数', '公共必修课通过门数', '专业必修课通过门数', '公选课通过门数',
                '公选课获得学分', '专选课通过门数', '专选课获得学分', '办公自动化', '综合技能训练通过门数(不含办公自动化)', '实践环节'):
        alldata[get_column_letter(col) + '1'] = tit
        result[get_column_letter(col) + '1'] = tit
        col += 1

    # 设置列宽
    alldata.column_dimensions['A'].width = 6
    alldata.column_dimensions['B'].width = 12
    alldata.column_dimensions['C'].width = 6
    alldata.column_dimensions['D'].width = 8
    alldata.column_dimensions['E'].width = 4
    alldata.column_dimensions['F'].width = 20
    alldata.column_dimensions['G'].width = 18
    alldata.column_dimensions['H'].width = 16

    result.column_dimensions['A'].width = 4
    result.column_dimensions['B'].width = 12
    result.column_dimensions['C'].width = 6
    result.column_dimensions['D'].width = 8
    result.column_dimensions['E'].width = 4
    result.column_dimensions['F'].width = 20
    result.column_dimensions['G'].width = 18
    result.column_dimensions['H'].width = 16

    # 定义单元格格式
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 上下左右居中,自动换行
    # 设置表头格式
    alldata['A1'].alignment = alignment
    alldata['B1'].alignment = alignment
    alldata['C1'].alignment = alignment
    alldata['D1'].alignment = alignment
    alldata['E1'].alignment = alignment
    alldata['F1'].alignment = alignment
    alldata['G1'].alignment = alignment
    alldata['H1'].alignment = alignment
    alldata['I1'].alignment = alignment
    alldata['J1'].alignment = alignment
    alldata['K1'].alignment = alignment
    alldata['L1'].alignment = alignment
    alldata['M1'].alignment = alignment
    alldata['N1'].alignment = alignment
    alldata['O1'].alignment = alignment
    alldata['P1'].alignment = alignment
    alldata['Q1'].alignment = alignment
    alldata['R1'].alignment = alignment
    alldata['S1'].alignment = alignment
    alldata['T1'].alignment = alignment
    alldata['U1'].alignment = alignment
    alldata['V1'].alignment = alignment

    result['A1'].alignment = alignment
    result['B1'].alignment = alignment
    result['C1'].alignment = alignment
    result['D1'].alignment = alignment
    result['E1'].alignment = alignment
    result['F1'].alignment = alignment
    result['G1'].alignment = alignment
    result['H1'].alignment = alignment
    result['I1'].alignment = alignment
    result['J1'].alignment = alignment
    result['K1'].alignment = alignment
    result['L1'].alignment = alignment
    result['M1'].alignment = alignment
    result['N1'].alignment = alignment
    result['O1'].alignment = alignment
    result['P1'].alignment = alignment
    result['Q1'].alignment = alignment
    result['R1'].alignment = alignment
    result['S1'].alignment = alignment
    result['T1'].alignment = alignment
    result['U1'].alignment = alignment
    result['V1'].alignment = alignment

    # 写入数据
    ro = 2
    # if zyxz == 0:
    #     for v, k in bys_re.items():
    #         if k[0][3] in xymc:
    #             alldata['A' + str(ro)] = ro - 1
    #             alldata['B' + str(ro)] = v
    #
    #             co = 3
    #             for ki in k[0]:
    #                 alldata[get_column_letter(co) + str(ro)] = ki
    #                 co += 1
    #
    #             coo = 9
    #             for kii in k[1].values():
    #                 alldata[get_column_letter(coo) + str(ro)] = kii
    #                 coo += 1
    #
    #             ro += 1

    # else:
    for v, k in bys_re.items():
        if k[0][4] in zymc:
            alldata['A' + str(ro)] = ro - 1
            alldata['B' + str(ro)] = v

            co = 3
            for ki in k[0]:
                alldata[get_column_letter(co) + str(ro)] = ki
                co += 1

            coo = 9
            for kii in k[1].values():
                alldata[get_column_letter(coo) + str(ro)] = kii
                coo += 1

            ro += 1

    # 工作表写保护
    wb.sheet = True
    alldata.protection.enable()
    alldata.protection.password = 'hqwy'

    '''
    遍历sheet(index=1, title='全部审核数据') 该sheet中所有数据，
    与用户输入的审核参数进行对比，
    如不满足，则将整行提取放到sheet1中，
    对不满足的字段标红
    '''
    # 用来保存各项目不合格的人数
    fail = {'jd': 0, 'gk': 0, 'tyk': 0, 'ggbxk': 0, 'zybxk': 0, 'gxk': 0, 'gxkxf': 0, 'zxk': 0, 'zxkxf': 0, 'bgzdh': 0,
            'zhjnxf': 0, 'sjhj': 0}

    font = Font(color='FFFF0000')

    one_row = 2
    for two_row in range(2, alldata.max_row + 1):
        # 绩点
        if float(alldata['I' + str(two_row)].value) < 2:
            # 整行复制
            for onetwo_column in range(2, alldata.max_column - 1):
                result['A' + str(one_row)] = one_row - 1  # A列重新编号
                result[get_column_letter(onetwo_column) + str(one_row)] = alldata[
                    get_column_letter(onetwo_column) + str(two_row)].value
                result['I' + str(one_row)].font = font
            fail['jd'] += 1
            one_row += 1

        # 挂科
        elif int(alldata['J' + str(two_row)].value) > 0:
            for onetwo_column in range(2, alldata.max_column - 1):
                result['A' + str(one_row)] = one_row - 1
                result[get_column_letter(onetwo_column) + str(one_row)] = alldata[
                    get_column_letter(onetwo_column) + str(two_row)].value
                # result['J' + str(one_row)].font = font
            one_row += 1
        # 体育
        elif int(alldata['K' + str(two_row)].value) < standard['tyk']:
            for onetwo_column in range(2, alldata.max_column - 1):
                result['A' + str(one_row)] = one_row - 1
                result[get_column_letter(onetwo_column) + str(one_row)] = alldata[
                    get_column_letter(onetwo_column) + str(two_row)].value
                # result['K' + str(one_row)].font = font
            one_row += 1
        # 公必门数
        elif int(alldata['L' + str(two_row)].value) < standard['ggbxk']:
            for onetwo_column in range(2, alldata.max_column - 1):
                result['A' + str(one_row)] = one_row - 1
                result[get_column_letter(onetwo_column) + str(one_row)] = alldata[
                    get_column_letter(onetwo_column) + str(two_row)].value
                # result['L' + str(one_row)].font = font
            one_row += 1
        # 专必门数
        elif int(alldata['M' + str(two_row)].value) < standard['zybxk']:
            for onetwo_column in range(2, alldata.max_column - 1):
                result['A' + str(one_row)] = one_row - 1
                result[get_column_letter(onetwo_column) + str(one_row)] = alldata[
                    get_column_letter(onetwo_column) + str(two_row)].value
                # result['M' + str(one_row)].font = font
            one_row += 1
        # 公选门数
        elif int(alldata['N' + str(two_row)].value) < standard['gxk']:
            for onetwo_column in range(2, alldata.max_column - 1):
                result['A' + str(one_row)] = one_row - 1
                result[get_column_letter(onetwo_column) + str(one_row)] = alldata[
                    get_column_letter(onetwo_column) + str(two_row)].value
                # result['N' + str(one_row)].font = font
            one_row += 1
        # 公选学分
        elif int(alldata['O' + str(two_row)].value) < standard['gxkxf']:
            for onetwo_column in range(2, alldata.max_column - 1):
                result['A' + str(one_row)] = one_row - 1
                result[get_column_letter(onetwo_column) + str(one_row)] = alldata[
                    get_column_letter(onetwo_column) + str(two_row)].value
                # result['O' + str(one_row)].font = font
            one_row += 1
        # 专选门数
        elif int(alldata['P' + str(two_row)].value) < standard['zxk']:
            for onetwo_column in range(2, alldata.max_column - 1):
                result['A' + str(one_row)] = one_row - 1
                result[get_column_letter(onetwo_column) + str(one_row)] = alldata[
                    get_column_letter(onetwo_column) + str(two_row)].value
                # result['P' + str(one_row)].font = font
            one_row += 1
        # 专选学分
        elif float(alldata['Q' + str(two_row)].value) < float(standard['zxkxf']):
            for onetwo_column in range(2, alldata.max_column - 1):
                result['A' + str(one_row)] = one_row - 1
                result[get_column_letter(onetwo_column) + str(one_row)] = alldata[
                    get_column_letter(onetwo_column) + str(two_row)].value
                # result['Q' + str(one_row)].font = font
            one_row += 1
        # 办公自动化
        elif int(alldata['R' + str(two_row)].value) < standard['bgzdh']:
            for onetwo_column in range(2, alldata.max_column - 1):
                result['A' + str(one_row)] = one_row - 1
                result[get_column_letter(onetwo_column) + str(one_row)] = alldata[
                    get_column_letter(onetwo_column) + str(two_row)].value
                # result['R' + str(one_row)].font = font
            one_row += 1
        # 综合技能
        elif int(alldata['S' + str(two_row)].value) < standard['zhjnxf']:
            for onetwo_column in range(2, alldata.max_column - 1):
                result['A' + str(one_row)] = one_row - 1
                result[get_column_letter(onetwo_column) + str(one_row)] = alldata[
                    get_column_letter(onetwo_column) + str(two_row)].value
                # result['S' + str(one_row)].font = font
            one_row += 1
        # 实践环节
        elif int(alldata['T' + str(two_row)].value) < standard['sjhj']:
            for onetwo_column in range(2, alldata.max_column - 1):
                result['A' + str(one_row)] = one_row - 1
                result[get_column_letter(onetwo_column) + str(one_row)] = alldata[
                    get_column_letter(onetwo_column) + str(two_row)].value
                # result['T' + str(one_row)].font = font
            one_row += 1

    # 遍历result表,把不合格的成绩表红
    standard_keys = ['jd', 'gk', 'tyk', 'ggbxk', 'zybxk', 'gxk', 'gxkxf', 'zxk', 'zxkxf', 'bgzdh', 'zhjnxf', 'sjhj']
    for one_roww in range(2, result.max_row + 1):
        if int(result['J' + str(one_roww)].value) > 0:
            result['J' + str(one_roww)].font = font
            fail['gk'] += 1
    for one_column in range(11, 21):
        for one_ro in range(2, result.max_row + 1):
            if float(result[get_column_letter(one_column) + str(one_ro)].value) < float(standard[
                                                                                            standard_keys[
                                                                                                one_column - 9]]):
                result[get_column_letter(one_column) + str(one_ro)].font = font
                fail[standard_keys[one_column - 9]] += 1

    # 不合格数据统计
    row_last = result.max_row  # 数据的最后一行
    result['H' + str(row_last + 1)] = '不合格人数统计：'
    column_last = 9
    for fail_value in fail.values():
        result[get_column_letter(column_last) + str(row_last + 1)] = fail_value
        column_last += 1

    # 毕业率
    result['H' + str(row_last + 2)] = '毕业率：'
    result['I' + str(row_last + 2)] = str(round(1 - (row_last - 1) / (alldata.max_row - 1), 5) * 100) + '%'

    # 格式
    font2 = Font(name=u'微软亚黑', color=colors.WHITE, bold=True)
    fill = PatternFill("solid", fgColor="71C671")
    result['H' + str(row_last + 2)].fill = fill  # 背景颜色
    result['I' + str(row_last + 2)].fill = fill
    result['H' + str(row_last + 2)].font = font2  # 字体
    result['I' + str(row_last + 2)].font = font2
    result['H' + str(row_last + 2)].alignment = Alignment(horizontal='center', vertical='center')  # 居中
    result['I' + str(row_last + 2)].alignment = Alignment(horizontal='right', vertical='center')

    # 边框
    # border1 = Border(top=Side(border_style=None, color='FF000000'))
    # border1 = Border(top=Side(border_style=None, color='FF000000'))
    # border2 = Border(bottom=Side(border_style=None, color='FF000000'))
    # border3 = Border(right=Side(border_style=None, color='FF000000'))
    # border4 = Border(left=Side(border_style=None, color='FF000000'))
    # for row_bor in range(1, row_last + 1):
    #     result['A' + str(row_bor)].border = border4

    # 绘制统计图
    projects = ['绩点', '挂科', '体育', '公必', '专必', '公选门数', '公选学分', '专选门数', '专选学分', '办公自动化', '综合技能', '实践环节']
    for i in range(2, 14):
        result[get_column_letter(i) + str(row_last + 4)] = projects[i - 2]
    yvalue = Reference(result, min_row=row_last + 1, min_col=8, max_col=20)
    xvalue = Reference(result, min_row=row_last + 4, min_col=2, max_col=13)

    chart1 = BarChart()
    chart1.add_data(data=yvalue, from_rows=True, titles_from_data=True)  # titles_from_data=True 将data区域的第一行作为名称
    chart1.set_categories(xvalue)

    chart1.type = "col"  # 纵向柱状， ='bar'为横向条状
    chart1.style = 36  # 样式，包括颜色
    chart1.title = "不合格人数统计图"
    # chart1.y_axis.title = '不合格人数'
    chart1.x_axis.title = '审核项目'
    chart1.width = 25
    chart1.height = 8
    result.add_chart(chart1, 'B' + str(row_last + 4))

    # 绘制绩点分布图
    # 绩点分段统计
    jdfd = {'绩点(含前不含后)': '人数', '<2': 0, '2--3': 0, '3--4': 0, '4--5': 0}
    jdfd['<2'] = fail['jd']
    for two_row2 in range(2, alldata.max_row + 1):
        if float(alldata['I' + str(two_row2)].value) >= 4:
            jdfd['4--5'] += 1
        elif float(alldata['I' + str(two_row2)].value) >= 3:
            jdfd['3--4'] += 1
        elif float(alldata['I' + str(two_row2)].value) >= 2:
            jdfd['2--3'] += 1

    ii = 14
    for key, val in jdfd.items():
        result[get_column_letter(ii) + str(row_last + 4)] = key
        result[get_column_letter(ii) + str(row_last + 5)] = val
        ii += 1

    yvalue2 = Reference(result, min_row=row_last + 5, min_col=14, max_col=18)
    xvalue2 = Reference(result, min_row=row_last + 4, min_col=15, max_col=18)

    chart2 = BarChart()
    chart2.add_data(data=yvalue2, from_rows=True, titles_from_data=True)  # titles_from_data=True 将data区域的第一行作为名称
    chart2.set_categories(xvalue2)

    chart2.type = "col"  # 纵向柱状， ='bar'为横向条状
    chart2.style = 36  # 样式，包括颜色
    chart2.title = "绩点段人数分布图"
    chart2.x_axis.title = '绩点(含前不含后)'
    chart2.width = 15.5
    chart2.height = 8
    result.add_chart(chart2, 'L' + str(row_last + 4))

    # 保存EXCEL，并判断重名，重名则后面加(n)
    # 如果学院选择了全校，则命名为“全校xx年毕业资格审核结果.xlsx"，否则命名为"学院、学院xx年毕业资格审核结果.xlsx"

    # 优化后方法
    if xyxz == 0:
        xymcc = '全校'
    else:
        xymcc = '、'.join(xymc)
    if (xymcc + str(j) + '届毕业资格审核结果.xlsx') not in os.listdir('.'):  # 如果本地目录下有重名文件，则在文件名后面加"(n)"，n顺次加1
        gra_fname = xymcc + str(j) + '届毕业资格审核结果.xlsx'
    else:
        fnum = 1
        while True:
            fnamee = xymcc + str(j) + '届毕业资格审核结果' + '(' + str(fnum) + ')' + '.xlsx'
            if fnamee not in os.listdir('.'):
                gra_fname = fnamee
                break
            else:
                fnum += 1
    wb.save(filename=gra_fname)

    # print('请查看文件《' + gra_fname + '》')
    aud_crl_otext.configure(text='请查看文件《' + gra_fname + '》\n并对不合格学生及下列问题数据进行复核', font=g_font)
    aud_crl_otext.pack()

    aud_crl_sh.configure(state='normal')


def audit_lookforError(conn):
    """毕业资格审核-将成绩中有cxbj=1并且重复kcdm的学生筛选出来，放在list Error_data中，没有连接数据库的步骤"""
    sql_lferroe = 'SELECT XH,XY,ZYMC FROM XSJBXXB WHERE XH IN ( SELECT XH FROM GRA_CJB WHERE XH||SUBSTR(XKKH,15,8) IN  (SELECT XH||SUBSTR(XKKH,15,8) FROM GRA_CJB WHERE CXBJ=1 ) AND CXBJ<>1)'
    cur = conn.cursor()
    cur.execute(sql_lferroe)
    data_list = [xh for xh in cur]
    return data_list


def connDB():
    """有课教师-连接数据库"""
    global kb, max_jsz, sub_control2_qszbox, sub_control2_jszbox, sub_control3_enq
    conn = cx_Oracle.connect('zfxfzb/zfsoft_hqwy@orcl')
    cur = conn.cursor()
    sql_kb = 'SELECT * from kb_allinfor'
    cur.execute(sql_kb)
    for row in cur:
        kb.append(row)
    cur.close()

    # 当学期最大结束周
    cur = conn.cursor()
    sql_kb = 'SELECT max(jsz) from kb_allinfor'
    cur.execute(sql_kb)
    max_jsz = [i for i in cur][0][0]
    cur.close()

    sub_control2_qszbox['values'] = list(range(1, max_jsz + 1))
    sub_control2_jszbox['values'] = list(range(1, max_jsz + 1))

    # print('连接成功', end='\n')
    sub_control1.configure(text='连接成功')  # 设置button显示的内容
    sub_control1.configure(state='disabled')  # 将按钮设置为灰色状态，不可使用状态
    sub_control3_enq.configure(state='normal')


def getnum():
    """有课教师-查询有课教师数量"""
    global kb, result, teachercount, sub_control3_enqtext, sub_control3_enqop
    input = [sub_control2_xqbox.get(), sub_control2_kjbox.get(), sub_control2_qszbox.get(), sub_control2_jszbox.get(),
             sub_control2_dszbox.get()]
    prar = {}
    if input[0] == '1-5':
        prar['xqj'] = [1, 2, 3, 4, 5]
    elif input[0] == '1-6' or input[0] == '':
        prar['xqj'] = [1, 2, 3, 4, 5, 6]
    else:
        prar['xqj'] = [int(input[0])]
    if input[1] == '白天' or input[1] == '':
        prar['sjd'] = ['1', '3', '5', '7']
    elif input[1] == '上午':
        prar['sjd'] = ['1', '3']
    elif input[1] == '下午':
        prar['sjd'] = ['5', '7']
    else:
        prar['sjd'] = [input[1][0]]

    if input[2] == '':
        prar['qsz'] = 20
    else:
        prar['qsz'] = int(input[2])
    # prar['jsz'] = int(input[3])
    if input[3] == '':
        prar['jsz'] = 1
    else:
        prar['jsz'] = int(input[3])

    if input[4] == '单周':
        prar['dsz'] = [None, '单']
    elif input[4] == '双周':
        prar['dsz'] = [None, '双']
    else:
        prar['dsz'] = [None, '单', '双']

    result = []
    teachercount = []

    for i in kb:
        if i[3] in prar['xqj'] and str(i[4]) in prar['sjd'] and i[5] in prar['dsz'] and i[6] <= prar['qsz'] and i[7] >= \
                prar['jsz']:
            result.append(i)
            teachercount.append(i[2])

    sub_control3_enqtext.configure(state='normal')
    sub_control3_enqtext.configure(text='该时间段有课教师有: ' + str(len(set(teachercount))) + '人', font=g_font)
    sub_control3_enqop.configure(state='normal')


def outtoexcel():
    """有课教师-输出到excel"""
    global result, out_end_text
    wb = openpyxl.Workbook()  # 打开新的工作簿
    table = wb.worksheets[0]
    table.title = '有课教师信息'

    # 写入表头
    col = 1
    for tit in ('序号', '教师', '课程名称', '星期几', '第几节', '地点', '单双周', '起始结束周', '课程性质', '开课学院'):
        table[get_column_letter(col) + '1'] = tit
        col += 1

    # 设置格式
    table.column_dimensions['A'].width = 7
    table.column_dimensions['C'].width = 36
    table.column_dimensions['F'].width = 10
    table.column_dimensions['H'].width = 12
    table.column_dimensions['I'].width = 14
    table.column_dimensions['J'].width = 18

    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 上下左右居中,自动换行
    for column in range(1, 11):
        table[get_column_letter(column) + '1'].alignment = alignment

    # 写入数据
    # ('2018-2019', '2', '13022', 2, 3, None, 10, 14, 2, '(2018-2019-2)-00014718-13022-1', '2002005014', 'qt', '中心-303',
    #  '姜丽媛', 'ERP业务实训2', '综合技能训练', '国际经济贸易学院')
    rownum = 2
    for row in result:
        table['A' + str(rownum)] = rownum - 1
        table['B' + str(rownum)] = row[13]
        table['C' + str(rownum)] = row[14]
        table['D' + str(rownum)] = '星期' + str(row[3])
        if row[8] == 1:
            table['E' + str(rownum)] = str(row[4]) + '节'
        else:
            table['E' + str(rownum)] = str(row[4]) + '、' + str(row[4] + 1) + '节'
        table['F' + str(rownum)] = row[12]
        table['G' + str(rownum)] = row[5]
        table['H' + str(rownum)] = str(row[6]) + '-' + str(row[7]) + '周'
        table['I' + str(rownum)] = row[-2]
        table['J' + str(rownum)] = row[-1]
        rownum += 1

    # 保存excel
    if ('有课教师信息.xlsx') not in os.listdir('.'):  # 如果本地目录下有重名文件，则在文件名后面加"(n)"，n顺次加1
        fname = '有课教师信息.xlsx'
    else:
        fnum = 1
        while True:
            fnamee = '有课教师信息' + '(' + str(fnum) + ')' + '.xlsx'
            if fnamee not in os.listdir('.'):
                fname = fnamee
                break
            else:
                fnum += 1
    wb.save(filename=fname)

    out_end_text.configure(text='输出完成，请查看: \n' + '《' + fname + '》', font=g_font)
    out_end_text.pack(fill='both', expand=1, padx=2, pady=2, side=tk.BOTTOM)


def grainfo_connDB():
    """毕业证/学位证号查询-连接数据库"""
    global gra_info, grainfo_crl1, grainfo_yeart
    gra_info = []
    conn = cx_Oracle.connect('zfxfzb/zfsoft_hqwy@orcl')
    cur = conn.cursor()
    sql_kb = 'SELECT XH,XM,XZB,RXRQ,SFZH,YWXM,YWCSD,XB,XY,ZYMC,DQSZJ,BDH FROM GRA_INFO'
    cur.execute(sql_kb)
    for row in cur:
        gra_info.append(row)
    cur.close()
    conn.close()

    grainfo_yeart.insert(-1, datetime.datetime.now().year)  # 设置Entry控件的默认值
    grainfo_crl1.configure(text='连接成功')  # 设置button显示的内容
    grainfo_crl1.configure(state='disabled')  # 将按钮设置为灰色状态，不可使用状态
    grainfo_bu.configure(state='normal')
    grainfo_but.configure(state='normal')


def gra_search(ev=None):
    """毕业证/学位证号查询-开始查询"""
    global grainfo_termsc, grainfo_termse, gra_info, grainfo_box3, resultForms
    grainfo_box3.pack(fill="both", expand=1, padx=2, side=tk.BOTTOM)
    s_result = []
    if grainfo_termsc.get() == '学号':
        for stu in gra_info:
            if stu[0] == grainfo_termse.get():
                s_result.append(stu[:7])
    elif grainfo_termsc.get() == '姓名':
        for stu in gra_info:
            if stu[1] == grainfo_termse.get():
                s_result.append(stu[:7])
    elif grainfo_termsc.get() == '身份证号':
        for stu in gra_info:
            if stu[8] == grainfo_termse.get():
                s_result.append(stu[:7])

    # 清空Treeview中的内容
    x = resultForms.get_children()
    for item in x:
        resultForms.delete(item)

    # 将新查询的数据放入Treeview
    for i in s_result:
        resultForms.insert('', i[0], text=i[0], values=i[0:])


def create_grainfo():
    """毕业生信息查询-生成毕业生的毕业证/学位证号码"""
    global grainfo_rulesc, grainfo_yesrt
    '''
    将grainfo_yesrt中选定毕业时间的毕业生插入table gra_info中，并将新数据按grainfo_rulesc中的规则生成毕业证和学位证
    '''
    pass


# 2、切换界面函数，选择功能后隐藏其他界面，显示自己界面
def btn_click_0(event=None):
    """用于功能选项卡的界面切换"""
    global frm_content_1
    btn_text = event.widget['text']
    if btn_text == '毕业资格审核':
        tch_box0.pack_forget()
        signupinfo_box0.pack_forget()
        grainfo_box0.pack_forget()
        audit_box0.pack(fill="both", expand=1, padx=2, side=tk.BOTTOM)

    elif btn_text == '有课教师查询':
        audit_box0.pack_forget()
        signupinfo_box0.pack_forget()
        grainfo_box0.pack_forget()
        tch_box0.pack(fill="both", expand=1, padx=2, side=tk.BOTTOM)

    elif btn_text == '报名信息查询':
        audit_box0.pack_forget()
        grainfo_box0.pack_forget()
        tch_box0.pack_forget()
        signupinfo_box0.pack(fill="both", expand=1, padx=2, side=tk.BOTTOM)


    elif btn_text == '毕业生信息查询':
        audit_box0.pack_forget()
        signupinfo_box0.pack_forget()
        tch_box0.pack_forget()
        grainfo_box0.pack(fill="both", expand=1, padx=2, side=tk.BOTTOM)


# 3、基本界面布局
root = tk.Tk()
root.title('吉林外国语大学——教务管理系统补充程序')
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
# root.iconbitmap(default=r'C:\Users\Administrator\Desktop\dogat.ico')
# 4、底层控件frm
frm = tk.Frame(root)
frm.pack(fill="both", expand=1)

# 5、菜单控件布局
frm_menu = tk.LabelFrame(frm)
frm_menu.pack(fill="both", expand=1, padx=2, side=tk.TOP)

menu_list = ['毕业资格审核', '有课教师查询', '报名信息查询', '毕业生信息查询']
for index, value in enumerate(menu_list):
    menu_button = tk.Button(frm_menu, anchor="w", text=value, font=g_font)
    menu_button.bind('<Button-1>', btn_click_0)
    menu_button.pack(fill="both", expand=1, padx=2, pady=2, side=tk.LEFT)

# 6、功能界面布局
audit_box0 = tk.LabelFrame(frm)  # 功能界面底层容器-毕业资格审核
audit_box0.pack(fill="both", expand=1, padx=2, side=tk.BOTTOM)

audit_box1 = tk.LabelFrame(audit_box0)
audit_box1.pack(fill="both", expand=1, padx=2, side=tk.TOP)
aud_crl_conn = tk.Button(audit_box0, text='点我连接数据库', font=g_font, command=audit_conn)
aud_crl_conn.pack(fill='both', expand=1, padx=2, pady=2)

audit_box2 = tk.LabelFrame(audit_box0, text='选择审核范围', font=g_font)
audit_box2.pack(fill="both", expand=1, padx=2, pady=10)
aud_crl_xy = tk.Label(audit_box2, text='学院：', font=g_font)
xy = tk.StringVar()
aud_crl_xyc = ttk.Combobox(audit_box2, textvariable=xy)
aud_crl_zy = tk.Label(audit_box2, text='  专业：', font=g_font)
zy = tk.StringVar()
aud_crl_zyc = ttk.Combobox(audit_box2, textvariable=zy)
aud_crl_zyc.bind('<Button-1>', audit_choosezy)
aud_crl_xy.pack(fill="both", expand=1, padx=2, pady=5, side=tk.LEFT)
aud_crl_xyc.pack(fill="both", expand=1, padx=2, pady=5, side=tk.LEFT)
aud_crl_zy.pack(fill="both", expand=1, padx=2, pady=5, side=tk.LEFT)
aud_crl_zyc.pack(fill="both", expand=1, padx=2, pady=5, side=tk.LEFT)

audit_box3 = tk.LabelFrame(audit_box0, text='填写审核标准', font=g_font)
audit_box3.pack(fill="both", expand=1, padx=2, pady=10)
audit_box31 = tk.Frame(audit_box3)
audit_box31.pack(fill="both", expand=1, padx=2, pady=10)
audit_box32 = tk.Frame(audit_box3)
audit_box32.pack(fill="both", expand=1, padx=2, pady=5)

aud_crl_gb = tk.LabelFrame(audit_box31, text='公共必修课', font=s_font)
aud_crl_gbs = tk.Scale(aud_crl_gb, from_=20, to=35, orient=tk.HORIZONTAL)
aud_crl_zb = tk.LabelFrame(audit_box31, text='专业必修课', font=s_font)
aud_crl_zbs = tk.Scale(aud_crl_zb, from_=15, to=42, orient=tk.HORIZONTAL)
aud_crl_tx = tk.LabelFrame(audit_box31, text='通选课门数', font=s_font)
aud_crl_txs = tk.Scale(aud_crl_tx, from_=0, to=4, orient=tk.HORIZONTAL)
aud_crl_txxf = tk.LabelFrame(audit_box31, text='通选课学分', font=s_font)
aud_crl_txxfs = tk.Scale(aud_crl_txxf, from_=0, to=6, orient=tk.HORIZONTAL)

aud_crl_zx = tk.LabelFrame(audit_box32, text='专选课门数', font=s_font)
aud_crl_zxs = tk.Scale(aud_crl_zx, from_=8, to=20, orient=tk.HORIZONTAL)
aud_crl_zxxf = tk.LabelFrame(audit_box32, text='专选课学分', font=s_font)
# aud_crl_zxxfs = tk.Scale(aud_crl_zxxf, from_=8, to=20, orient=tk.HORIZONTAL)
varr = tk.StringVar()
aud_crl_zxxfe = tk.Entry(aud_crl_zxxf, bd=2, textvariable=varr)
aud_crl_zj = tk.LabelFrame(audit_box32, text='综合技能训练', font=s_font)
aud_crl_zjs = tk.Scale(aud_crl_zj, from_=0, to=6, orient=tk.HORIZONTAL)
aud_crl_sj = tk.LabelFrame(audit_box32, text='实践环节', font=s_font)
aud_crl_sjs = tk.Scale(aud_crl_sj, from_=0, to=3, orient=tk.HORIZONTAL)

aud_crl_gb.pack(padx=2, side=tk.LEFT)
aud_crl_gbs.pack(padx=2, side=tk.LEFT)
aud_crl_zb.pack(padx=2, side=tk.LEFT)
aud_crl_zbs.pack(padx=2, side=tk.LEFT)
aud_crl_zj.pack(padx=2, side=tk.LEFT)
aud_crl_zjs.pack(padx=2, side=tk.LEFT)
aud_crl_sj.pack(padx=2, side=tk.LEFT)
aud_crl_sjs.pack(padx=2, side=tk.LEFT)
aud_crl_zx.pack(padx=2, side=tk.LEFT)
aud_crl_zxs.pack(padx=2, side=tk.LEFT)
aud_crl_zxxf.pack(padx=2, side=tk.LEFT)
aud_crl_zxxfe.pack(padx=2, pady=10, side=tk.LEFT)
aud_crl_tx.pack(padx=2, side=tk.LEFT)
aud_crl_txs.pack(padx=2, side=tk.LEFT)
aud_crl_txxf.pack(padx=2, side=tk.LEFT)
aud_crl_txxfs.pack(padx=2, side=tk.LEFT)

audit_box4 = tk.Frame(audit_box0)
audit_box4.pack(fill="both", expand=1, padx=2, pady=10)
var_sh = tk.StringVar()
var_sh.set('开始审核')
aud_crl_sh = tk.Button(audit_box4, textvariable=var_sh, font=g_font, )
aud_crl_sh.bind('<ButtonRelease-1>', begin_audit)
# aud_crl_sh.bind('<ButtonRelease-1>', begin_audit)
# aud_crl_sh.bind('<ButtonRelease-1>', begin_audit2)
aud_crl_sc = tk.Button(audit_box4, text='输出到Excel', font=g_font, command=audit_toexcel)
aud_crl_otext = tk.Label(audit_box4, font=g_font)
aud_crl_sh.pack(fill='both', padx=2, pady=1)
aud_crl_sc.pack(fill='both', padx=2, pady=5)
aud_crl_sh.configure(state='disable')  # 暂时设为灰色，查询成功后显示
aud_crl_sc.configure(state='disable')  # 暂时设为灰色，查询成功后显示

audit_box5 = tk.LabelFrame(audit_box0, text='有问题的数据（请根据成绩总表进行复核）：', font=g_font)
audit_box5.pack(fill="both", expand=1, padx=2)
aud_crl_treev = ttk.Treeview(audit_box5, height=5, show="headings", columns=('xy', 'xh', 'xm', 'xzb', 'jd', 'gk'))
"""点击字段排序"""
for col in ('xy', 'xh', 'xm', 'xzb', 'jd', 'gk'):
    aud_crl_treev.heading(col, text=col,
                          command=lambda c=col: treeview_sort_column(aud_crl_treev, c, False))

aud_crl_treev.pack(side=tk.LEFT)
aud_crl_treev.heading('xh', text='学号')  # 设置字段的显示名称
aud_crl_treev.heading('xm', text='姓名')
aud_crl_treev.heading('xy', text='学院')
aud_crl_treev.heading('xzb', text='班级')
aud_crl_treev.heading('jd', text='绩点')
aud_crl_treev.heading('gk', text='挂科数')

aud_crl_treev.column('xy', width=100)
aud_crl_treev.column('xh', width=80)  # 设置字段宽度
aud_crl_treev.column('xm', width=80)
aud_crl_treev.column('xzb', width=80)
aud_crl_treev.column('jd', width=80)
aud_crl_treev.column('gk', width=60)

aud_crl_sbar = ttk.Scrollbar(audit_box5, orient=tk.VERTICAL, command=aud_crl_treev.yview)
aud_crl_treev.configure(yscrollcommand=aud_crl_sbar.set)
aud_crl_sbar.pack(side=tk.LEFT, fill='y')

audit_box0.pack_forget()

tch_box0 = tk.LabelFrame(frm)  # 功能界面底层容器-有课教师查询
tch_box0.pack(fill="both", expand=1, padx=2, side=tk.BOTTOM)
tch_text = tk.Label(tch_box0, text='本功能可按时间查询有课教师\n请按照1、2、3步完成操作。', font=g_font)
tch_text.pack(fill='both', expand=1, padx=2, pady=2, side=tk.TOP)

tch_box1 = tk.LabelFrame(tch_box0)
tch_box1.pack(fill="both", expand=1, padx=2)
sub_control1 = tk.Button(tch_box1, text='1、点我连接数据库', font=g_font, command=connDB)
sub_control1.pack(fill='both', expand=1, padx=2, pady=2)

tch_box2 = tk.LabelFrame(tch_box0)
tch_box2.pack(fill="both", expand=1, padx=2)
sub_control2 = tk.Label(tch_box2, text='2、选择时间段', font=g_font)
sub_control2.pack(fill="both", expand=1, padx=2)

tch_box21 = tk.Frame(tch_box2)
tch_box21.pack(fill="both", expand=1, padx=2)
sub_control2_xq = tk.Label(tch_box21, text='星期：', font=g_font)
sub_control2_xq.pack(padx=10, pady=2, side=tk.LEFT, expand='no')
number = tk.StringVar()
sub_control2_xqbox = ttk.Combobox(tch_box21, textvariable=number)
sub_control2_xqbox['values'] = (1, 2, 3, 4, 5, 6, '1-5', '1-6')
sub_control2_xqbox.pack(padx=2, pady=2, side=tk.LEFT)

sub_control2_kj = tk.Label(tch_box21, text='课节：', font=g_font)
sub_control2_kj.pack(padx=10, pady=2, side=tk.LEFT)
number = tk.StringVar()
sub_control2_kjbox = ttk.Combobox(tch_box21, textvariable=number)
sub_control2_kjbox['values'] = ('1、2节', '3、4节', '5、6节', '7、8节', '9、10节', '上午', '下午', '白天')
sub_control2_kjbox.pack(padx=2, pady=2, side=tk.LEFT)

tch_box22 = tk.Frame(tch_box2)
tch_box22.pack(fill="both", expand=1, padx=2)
sub_control2_qzz = tk.Label(tch_box22, text='周次：', font=g_font)
sub_control2_qzz.pack(fill='both', padx=10, pady=3, side=tk.LEFT)
number = tk.StringVar()
sub_control2_qszbox = ttk.Combobox(tch_box22, textvariable=number)
sub_control2_qszbox.pack(padx=2, pady=3, side=tk.LEFT)
sub_control2_qzz = tk.Label(tch_box22, text=' ——  ', font=g_font)
sub_control2_qzz.pack(padx=10, pady=3, side=tk.LEFT)
number = tk.StringVar()
sub_control2_jszbox = ttk.Combobox(tch_box22, textvariable=number)
sub_control2_jszbox.pack(padx=0, pady=3, side=tk.LEFT)

tch_box23 = tk.Frame(tch_box2)
tch_box23.pack(fill="both", expand=1, padx=2)
sub_control2_dsz = tk.Label(tch_box23, text='单双周：', font=g_font)
sub_control2_dsz.pack(fill='both', padx=2, pady=3, side=tk.LEFT)
sub_control2_dszbox = ttk.Combobox(tch_box23)
sub_control2_dszbox['values'] = ('', '单周', '双周')
sub_control2_dszbox.pack(fill='both', padx=2, pady=3, side=tk.LEFT)

tch_box3 = tk.LabelFrame(tch_box0)  # 查询结果控件的容器
tch_box3.pack(fill="both", expand=1, padx=2)
sub_control3_enq = tk.Button(tch_box3, text='3、点击查询', font=g_font, command=getnum)
sub_control3_enq.pack(fill='both', expand=1, padx=2, pady=3, side=tk.TOP)
sub_control3_enq.configure(state='disable')  # 暂时设为灰色，查询成功后显示
sub_control3_enqtext = tk.Label(tch_box3, text='查询结果', font=g_font)
sub_control3_enqtext.pack(fill='both', expand=1, padx=2, pady=2)
sub_control3_enqtext.configure(state='disable')  # 暂时设为灰色，查询成功后显示
sub_control3_enqop = tk.Button(tch_box3, text='详细名单输出到Excel', font=g_font, command=outtoexcel)
sub_control3_enqop.pack(fill='both', expand=1, padx=2, pady=3)
sub_control3_enqop.configure(state='disable')  # 暂时设为灰色，查询成功后显示
out_end_text = tk.Label(tch_box3)
tch_box0.pack_forget()

signupinfo_box0 = tk.LabelFrame(frm)  # 功能界面底层容器-报名信息查询
signupinfo_box0.pack(fill="both", expand=1, padx=2, side=tk.BOTTOM)
signup_crl1 = tk.Button(signupinfo_box0, text='1、点我连接数据库', font=g_font, command=connDB)
signup_crl1.pack(fill='both', expand=1, padx=2, pady=2)

signupinfo_box0.pack_forget()

grainfo_box0 = tk.LabelFrame(frm)  # 功能界面底层容器-毕业生信息查询
grainfo_box0.pack(fill="both", expand=1, padx=2, side=tk.BOTTOM)
grainfo_crl1 = tk.Button(grainfo_box0, text='点我连接数据库', font=g_font, command=grainfo_connDB)
grainfo_crl1.pack(fill='both', expand=1, padx=2, pady=2)

grainfo_box1 = tk.LabelFrame(grainfo_box0, text='生成毕业生信息', font=g_font)
grainfo_box1.pack(fill="both", expand=1, padx=2, pady=15)

grainfo_rules = tk.Label(grainfo_box1, text='生成规则：', font=g_font)
number = tk.StringVar()
grainfo_rulesc = ttk.Combobox(grainfo_box1, textvariable=number)
grainfo_rulesc['values'] = ('学号正序', '学号倒叙', '报到号正序', '报到号倒叙', '身份证号正序', '身份证号倒叙')
grainfo_year = tk.Label(grainfo_box1, text=' 毕业年份：', font=g_font)
grainfo_yeart = tk.Entry(grainfo_box1, bd=2)  # bd为边框大小，缺省值为1
grainfo_bu = tk.Button(grainfo_box1, text='生成', font=g_font, command=create_grainfo)
grainfo_rules.pack(expand=1, padx=2, pady=2, side=tk.LEFT)
grainfo_rulesc.pack(expand=1, padx=2, pady=2, side=tk.LEFT)
grainfo_year.pack(expand=1, padx=2, pady=2, side=tk.LEFT)
grainfo_yeart.pack(expand=1, padx=2, pady=2, side=tk.LEFT)
grainfo_bu.pack(expand=1, padx=10, pady=2, side=tk.LEFT)
grainfo_bu.configure(state='disable')

grainfo_box2 = tk.LabelFrame(grainfo_box0, text='毕业生信息查询', font=g_font)
grainfo_box2.pack(fill="both", expand=1, padx=2, pady=15)
grainfo_terms = tk.Label(grainfo_box2, text='查询条件：', font=g_font)
number = tk.StringVar()
grainfo_termsc = ttk.Combobox(grainfo_box2, textvariable=number)
grainfo_termsc['values'] = ('学号', '姓名', '身份证号')
var = tk.StringVar()
grainfo_in = tk.Label(grainfo_box2, text=' 输入内容：', font=g_font)
grainfo_termse = tk.Entry(grainfo_box2, bd=2, textvariable=var)
grainfo_termse.bind('<Return>', gra_search)
grainfo_but = tk.Button(grainfo_box2, text='查询', font=g_font, command=gra_search)
grainfo_terms.pack(expand=1, padx=2, pady=2, side=tk.LEFT)
grainfo_termsc.pack(expand=1, padx=2, pady=2, side=tk.LEFT)
grainfo_in.pack(expand=1, padx=2, pady=2, side=tk.LEFT)
grainfo_termse.pack(expand=1, padx=2, pady=2, side=tk.LEFT)
grainfo_but.pack(expand=1, padx=10, pady=2, side=tk.LEFT)
grainfo_but.configure(state='disable')

grainfo_box3 = tk.LabelFrame(grainfo_box0, text='查询结果', font=g_font)

# 创建Treeview来展示查询到的数据
resultForms = ttk.Treeview(grainfo_box3, height=5, show="headings",
                           columns=('xh', 'xm', 'xzb', 'rxrq', 'sfzh', 'byzh',
                                    'xwzh'))  # height控件高度，show=隐藏Treeview中的首列 column设置字段

"""点击字段排序"""
for col in ('xh', 'xm', 'xzb', 'rxrq', 'sfzh', 'byzh', 'xwzh'):
    resultForms.heading(col, text=col,
                        command=lambda c=col: treeview_sort_column(resultForms, c, False))

resultForms.pack(side=tk.LEFT)
resultForms.heading('xh', text='学号')  # 设置字段的显示名称
resultForms.heading('xm', text='姓名')
resultForms.heading('xzb', text='班级')
resultForms.heading('rxrq', text='入学日期')
resultForms.heading('sfzh', text='身份证号')
resultForms.heading('byzh', text='毕业证号')
resultForms.heading('xwzh', text='学位证号')
resultForms.column('xh', width=80)  # 设置字段宽度
resultForms.column('xm', width=60)
resultForms.column('xzb', width=80)
resultForms.column('rxrq', width=60)
resultForms.column('sfzh', width=120)
resultForms.column('byzh', width=120)
resultForms.column('xwzh', width=120)
vbar = ttk.Scrollbar(grainfo_box3, orient=tk.VERTICAL, command=resultForms.yview)  # 设置滚动条控件
vbar.pack(side=tk.LEFT, fill='y')  # 滚动条控件与被控制的Treeview同在一个容器中，并列放置，纵向填充
resultForms.configure(
    yscrollcommand=vbar.set)  # 此处被控控件的 yscrollcommand=vbar.set 与上面滚动条控件的 command=resultForms.yview 为相互作用设置
grainfo_box0.pack_forget()

root.mainloop()
